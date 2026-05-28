# backend/app/api/staff.py
import asyncio
import io
import json
import time
import openpyxl
from fastapi import APIRouter, Body, Depends, UploadFile, File, HTTPException
from app.models.customer import Customer
from app.models.conversation import Conversation
from app.models.product import Product
from app.models.order import Order
from app.api.deps import get_db, get_current_admin
from app.services.data_importer import parse_excel_columns, merge_product, import_products_from_wb

router = APIRouter(prefix="/api/staff", tags=["staff"], dependencies=[Depends(get_current_admin)])


@router.get("/products")
def list_products(
    page: int = 1,
    page_size: int = 50,
    category: str = None,
    stock_zero: bool = False,
    q: str = None,
    db=Depends(get_db)
):
    """搜索商品，支持按名称、商品条码、成分关键词搜索。"""
    stmt = db.query(Product)
    if category:
        stmt = stmt.filter(Product.category == category)
    if stock_zero:
        stmt = stmt.filter(Product.stock <= 0)
    if q:
        q_lower = q.lower()
        stmt = stmt.filter(
            (Product.name.ilike(f"%{q}%")) |
            (Product.sku_id.ilike(f"%{q}%")) |
            (Product.ingredients.ilike(f"%{q}%")) |
            (Product.category.ilike(f"%{q}%"))
        )
    total = stmt.count()
    offset = (page - 1) * page_size
    products = stmt.order_by(Product.name).offset(offset).limit(page_size).all()
    return {
        "total": total,
        "products": [
            {"id": p.id, "sku_id": p.sku_id, "name": p.name, "category": p.category,
             "price": p.price, "stock": p.stock, "is_active": p.is_active,
             "ingredients": p.ingredients,
             "scene_tags": p.scene_tags or "", "contraindication_tags": p.contraindication_tags or ""}
            for p in products
        ]
    }


@router.get("/products/by-skus")
def get_products_by_skus(skus: str, db=Depends(get_db)):
    """根据逗号分隔的 sku_id 列表查询商品"""
    sku_list = [s.strip() for s in skus.split(",") if s.strip()]
    products = db.query(Product).filter(Product.sku_id.in_(sku_list)).all()
    return {
        "products": [
            {"id": p.id, "sku_id": p.sku_id, "name": p.name, "category": p.category,
             "price": p.price, "stock": p.stock, "is_active": p.is_active,
             "scene_tags": p.scene_tags or "", "contraindication_tags": p.contraindication_tags or ""}
            for p in products
        ]
    }


@router.patch("/products/batch-stock-relative")
def batch_update_stock_relative(updates: list[dict], db=Depends(get_db)):
    """批量相对更新库存，如 [{"sku_id": "P26...", "delta": 5}] 或 [{"sku_id": "P26...", "stock": 10}]"""
    updated = []
    for item in updates:
        sku = item.get("sku_id")
        if not sku:
            continue
        product = db.query(Product).filter(Product.sku_id == sku).first()
        if not product:
            continue
        delta = item.get("delta")
        stock = item.get("stock")
        if delta is not None:
            product.stock = max(0, (product.stock or 0) + int(delta))
        elif stock is not None:
            product.stock = max(0, int(stock))
        updated.append({"sku_id": sku, "stock": product.stock})
    db.commit()
    return {"updated": len(updated), "items": updated}


@router.patch("/products/{sku_id}")
def update_product(sku_id: str, data: dict = Body(...), db=Depends(get_db)):
    """更新单个商品的所有字段"""
    product = db.query(Product).filter(Product.sku_id == sku_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="商品不存在")
    updatable = ["name", "category", "price", "stock", "ingredients", "is_active",
                  "scene_tags", "contraindication_tags", "feature_tag", "sales_script"]
    for field in updatable:
        if field in data:
            val = data[field]
            if field == "price":
                val = float(val) if val is not None else 0
            elif field == "stock":
                val = int(val) if val is not None else 0
            elif field == "is_active":
                val = bool(val)
            setattr(product, field, val)
    db.commit()
    return {"status": "ok", "sku_id": sku_id}


@router.get("/categories")
def list_categories(db=Depends(get_db)):
    from sqlalchemy import func
    rows = db.query(Product.category, func.count(Product.id)).group_by(Product.category).all()
    return {
        "categories": [
            {"name": cat, "count": cnt}
            for cat, cnt in sorted(rows, key=lambda r: r[0] or "")
            if cat
        ]
    }


@router.post("/products/preview")
async def preview_import(file: UploadFile = File(...), db=Depends(get_db)):
    """上传 Excel，返回预览（映射+样例+警告），不写 DB。"""
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 文件")
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    field_map = parse_excel_columns(headers)

    total_rows = sum(1 for row in ws.iter_rows(min_row=2) if any(c.value for c in row))

    samples = []
    warnings = []
    existing_skus = {p.sku_id for p in db.query(Product.sku_id).all()}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=7, values_only=True), start=2):
        row_data = {}
        for field_name, col_idx in field_map.items():
            val = row[col_idx] if col_idx < len(row) else None
            row_data[field_name] = val

        sku = row_data.get("sku_id", "")
        name = row_data.get("name", "")

        is_garbled = False
        if name:
            try:
                name.encode("utf-8")
            except UnicodeEncodeError:
                is_garbled = True

        if is_garbled:
            warnings.append({
                "row": row_idx,
                "type": "garbled",
                "sku_id": sku or "(empty)",
                "message": "商品名乱码，请在有赞后台修正后重新导入"
            })
        elif sku and sku in existing_skus:
            warnings.append({
                "row": row_idx,
                "type": "duplicate",
                "sku_id": sku,
                "message": "数据库已存在，将更新主数据"
            })

        samples.append({k: v for k, v in row_data.items() if k})

    return {
        "total_rows": total_rows,
        "mapping": {v: k for k, v in field_map.items()},
        "warnings": warnings,
        "samples": samples
    }


@router.post("/products/import")
async def import_excel(file: UploadFile = File(...), db=Depends(get_db)):
    """上传有赞导出的商品库 Excel，增量导入（主数据用导入值，Tag 保留/补全）"""
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 文件")
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))

    result = import_products_from_wb(db, wb)
    return {
        "imported": result["imported"],
        "updated": result["updated"],
        "tag_filled": result["tag_filled"],
        "skipped_garbled": result["skipped_garbled"],
        "errors": []
    }


@router.post("/products/generate-tags")
async def generate_missing_tags(db=Depends(get_db)):
    """对所有 scene_tags 为空的有效产品批量生成标签"""
    products = db.query(Product).filter(
        Product.is_active == True,
        Product.stock > 0,
        (Product.scene_tags == None) | (Product.scene_tags == "")
    ).all()

    gen = TagGenerator()
    success = 0
    failed = []
    for p in products:
        try:
            tags = gen.generate(p.name, p.ingredients or "")
            p.scene_tags = tags["scene_tags"]
            p.contraindication_tags = tags["contraindication_tags"]
            db.commit()
            success += 1
        except Exception as e:
            await asyncio.sleep(5)
            try:
                tags = gen.generate(p.name, p.ingredients or "")
                p.scene_tags = tags["scene_tags"]
                p.contraindication_tags = tags["contraindication_tags"]
                db.commit()
                success += 1
            except Exception as e2:
                failed.append({"sku_id": p.sku_id, "name": p.name, "reason": str(e2)[:100]})
                if len(failed) % 10 == 0:
                    db.commit()

    db.commit()
    return {"total": len(products), "success": success, "failed": failed}


@router.get("/customers")
def list_customers(page: int = 1, page_size: int = 20, db=Depends(get_db)):
    offset = (page - 1) * page_size
    customers = db.query(Customer).order_by(Customer.created_at.desc()).offset(offset).limit(page_size).all()
    return {"customers": [
        {"id": c.id, "nickname": c.nickname, "phone": c.phone,
         "constitution": c.constitution, "created_at": str(c.created_at)}
        for c in customers
    ]}


@router.get("/customers/{customer_id}")
def customer_detail(customer_id: int, db=Depends(get_db)):
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    conversations = db.query(Conversation).filter(Conversation.customer_id == customer_id).all()
    return {"customer": customer, "conversations": conversations}


# ── Orders ──────────────────────────────────────────────────────────────────

def _order_dict(o: Order) -> dict:
    """Serialize an Order row into a clean JSON-safe dict."""
    return {
        "id": o.id,
        "order_no": o.order_no,
        "customer_id": o.customer_id,
        "customer_nickname": o.customer_nickname,
        "customer_phone": o.customer_phone,
        "total_amount": o.total_amount,
        "status": o.status,
        "items": json.loads(o.items_json) if o.items_json else [],
        "conversation_snapshot": json.loads(o.conversation_snapshot) if o.conversation_snapshot else None,
        "recommendation_snapshot": json.loads(o.recommendation_snapshot) if o.recommendation_snapshot else None,
        "constitution_snapshot": json.loads(o.constitution_snapshot) if o.constitution_snapshot else None,
        "created_at": str(o.created_at) if o.created_at else None,
        "paid_at": str(o.paid_at) if o.paid_at else None,
    }


@router.post("/orders")
def create_order(data: dict, db=Depends(get_db)):
    """Create/save an order when customer completes checkout in mini-program."""
    existing = db.query(Order).filter(Order.order_no == data.get("order_no")).first()
    if existing:
        raise HTTPException(status_code=409, detail="订单号已存在")
    order = Order(
        order_no=data["order_no"],
        customer_nickname=data.get("customer_nickname"),
        customer_phone=data.get("customer_phone"),
        total_amount=float(data.get("total_amount", 0)),
        items_json=json.dumps(data.get("items", []), ensure_ascii=False),
        conversation_snapshot=json.dumps(data.get("conversation_snapshot"), ensure_ascii=False) if data.get("conversation_snapshot") else None,
        recommendation_snapshot=json.dumps(data.get("recommendation_snapshot"), ensure_ascii=False) if data.get("recommendation_snapshot") else None,
        constitution_snapshot=json.dumps(data.get("constitution_snapshot"), ensure_ascii=False) if data.get("constitution_snapshot") else None,
    )
    db.add(order)
    db.commit()
    db.refresh(order)
    return _order_dict(order)


@router.get("/orders")
def list_orders(page: int = 1, page_size: int = 20, status: str = None, db=Depends(get_db)):
    """List all orders, newest first. Status filter: pending/paid/cancelled/refunded."""
    stmt = db.query(Order)
    if status:
        stmt = stmt.filter(Order.status == status)
    total = stmt.count()
    offset = (page - 1) * page_size
    orders = stmt.order_by(Order.created_at.desc()).offset(offset).limit(page_size).all()
    return {"total": total, "orders": [_order_dict(o) for o in orders]}


@router.get("/orders/{order_id}")
def order_detail(order_id: int, db=Depends(get_db)):
    """Get order detail with full conversation/recommendation/constitution snapshots."""
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="订单不存在")
    return _order_dict(order)


@router.patch("/orders/{order_id}/status")
def update_order_status(order_id: int, status: str, db=Depends(get_db)):
    """Update order status: paid/cancelled/refunded"""
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="订单不存在")
    valid = ("paid", "cancelled", "refunded")
    if status not in valid:
        raise HTTPException(status_code=400, detail=f"状态必须是: {', '.join(valid)}")
    order.status = status
    if status == "paid":
        from datetime import datetime
        order.paid_at = datetime.utcnow()
    db.commit()
    return _order_dict(order)


@router.get("/constitution-bundles")
def get_constitution_bundles(db=Depends(get_db)):
    """获取所有体质套餐，按体质类型分组"""
    from app.models.constitution_bundle import ConstitutionBundle
    from app.models.product import Product
    bundles = db.query(ConstitutionBundle).order_by(ConstitutionBundle.sort_order).all()
    grouped = {}
    for b in bundles:
        prod = db.query(Product).filter_by(sku_id=b.sku_id).first()
        grouped.setdefault(b.constitution_type, []).append({
            "sku_id": b.sku_id,
            "sort_order": b.sort_order,
            "description": b.description,
            "name": prod.name if prod else b.sku_id,
            "category": prod.category if prod else "—",
        })
    return grouped


@router.put("/constitution-bundles/{ctype}")
def put_constitution_bundle(ctype: str, data: dict = Body(...), db=Depends(get_db)):
    """整体替换某体质类型的套餐"""
    from app.models.constitution_bundle import ConstitutionBundle
    db.query(ConstitutionBundle).filter(ConstitutionBundle.constitution_type == ctype).delete()
    db.flush()
    for i, p in enumerate(data.get("products", [])):
        item = ConstitutionBundle(
            constitution_type=ctype,
            sku_id=p["sku_id"],
            sort_order=p.get("sort_order", i),
            description=p.get("description", "")
        )
        db.add(item)
    db.commit()
    return {"status": "ok", "count": len(data.get("products", []))}


@router.get("/hot-products")
def get_hot_products(db=Depends(get_db)):
    """获取主推产品列表"""
    from app.models.constitution_bundle import HotProduct
    from app.models.product import Product
    items = db.query(HotProduct).order_by(HotProduct.sort_order).all()
    # Pre-fetch all needed products in one query to avoid N+1
    sku_ids = [h.sku_id for h in items]
    product_map = {p.sku_id: p for p in db.query(Product).filter(Product.sku_id.in_(sku_ids)).all()}
    return [
        {"sku_id": h.sku_id, "sort_order": h.sort_order,
         "name": product_map[h.sku_id].name if h.sku_id in product_map else h.sku_id,
         "category": product_map[h.sku_id].category if h.sku_id in product_map else "—"}
        for h in items
    ]


@router.put("/hot-products")
def put_hot_products(data: dict = Body(...), db=Depends(get_db)):
    """整体替换主推产品列表"""
    from app.models.constitution_bundle import HotProduct
    db.query(HotProduct).delete()
    db.flush()
    for i, p in enumerate(data.get("products", [])):
        item = HotProduct(sku_id=p["sku_id"], sort_order=p.get("sort_order", i))
        db.add(item)
    db.commit()
    return {"status": "ok", "count": len(data.get("products", []))}


# ── Channel tracking ────────────────────────────────────────────────────────

@router.get("/channels")
def list_channels(db=Depends(get_db)):
    """按渠道统计对话数量和转化"""
    from sqlalchemy import func as sqlfunc
    rows = db.query(
        Conversation.channel,
        sqlfunc.count(Conversation.id)
    ).group_by(Conversation.channel).all()
    channels = []
    for channel, count in rows:
        ch = channel or "直接"
        total_orders = db.query(Order).filter(
            Order.conversation_snapshot.isnot(None)
        ).count() if ch != "直接" else 0
        channels.append({
            "channel": ch,
            "conversation_count": count,
            "order_count": 0,  # requires order-to-conversation linking
        })
    return {"channels": channels}


# ── Conversation list & staff marking ───────────────────────────────────────

@router.get("/conversations")
def list_conversations(
    page: int = 1,
    page_size: int = 20,
    flagged_only: bool = False,
    channel: str = None,
    db=Depends(get_db)
):
    """列出对话记录，支持按标记和渠道筛选"""
    stmt = db.query(Conversation)
    if flagged_only:
        stmt = stmt.filter(Conversation.is_flagged == True)
    if channel:
        stmt = stmt.filter(Conversation.channel == channel)
    total = stmt.count()
    offset = (page - 1) * page_size
    convs = stmt.order_by(Conversation.created_at.desc()).offset(offset).limit(page_size).all()
    return {
        "total": total,
        "conversations": [
            {
                "id": c.id,
                "customer_id": c.customer_id,
                "stage": c.stage,
                "messages": c.messages,
                "messages_history": json.loads(c.messages_history) if c.messages_history else [],
                "stage_history": json.loads(c.stage_history) if c.stage_history else [],
                "channel": c.channel or "",
                "screening_result": c.screening_result,
                "staff_notes": c.staff_notes or "",
                "staff_tags": c.staff_tags or "",
                "is_flagged": c.is_flagged,
                "created_at": str(c.created_at),
            }
            for c in convs
        ]
    }


@router.get("/conversations/{conv_id}")
def get_conversation(conv_id: int, db=Depends(get_db)):
    """获取单条对话详情"""
    conv = db.query(Conversation).filter(Conversation.id == conv_id).first()
    if not conv:
        raise HTTPException(status_code=404, detail="对话不存在")
    return {
        "id": conv.id,
        "customer_id": conv.customer_id,
        "stage": conv.stage,
        "messages": conv.messages,
        "messages_history": json.loads(conv.messages_history) if conv.messages_history else [],
        "stage_history": json.loads(conv.stage_history) if conv.stage_history else [],
        "channel": conv.channel or "",
        "screening_result": conv.screening_result,
        "constitution_type": conv.constitution_override or conv.constitution_type or "",
        "constitution_ai": conv.constitution_type or "",
        "constitution_confidence": conv.constitution_confidence or 0,
        "staff_notes": conv.staff_notes or "",
        "staff_tags": conv.staff_tags or "",
        "is_flagged": conv.is_flagged,
        "created_at": str(conv.created_at),
    }


@router.patch("/conversations/{conv_id}/constitution")
def override_constitution(conv_id: int, data: dict = Body(...), db=Depends(get_db)):
    """店员手动修正体质类型"""
    conv = db.query(Conversation).filter(Conversation.id == conv_id).first()
    if not conv:
        raise HTTPException(status_code=404, detail="对话不存在")
    ctype = data.get("constitution_type", "")
    if ctype:
        conv.constitution_override = ctype
    db.commit()
    return {
        "id": conv_id,
        "constitution_type": conv.constitution_type,
        "constitution_override": conv.constitution_override,
    }


@router.get("/reports/constitution-performance")
def constitution_performance(db=Depends(get_db)):
    """各体质转化率：对话数 → 订单数 → 转化率"""
    from sqlalchemy import func as sqlfunc
    # Count conversations by constitution (AI or override)
    rows = db.query(
        Conversation.constitution_type,
        sqlfunc.count(Conversation.id)
    ).filter(Conversation.constitution_type.isnot(None)).group_by(
        Conversation.constitution_type
    ).all()

    result = []
    for ctype, count in rows:
        # Count orders for customers with this constitution
        cust_ids = db.query(Conversation.customer_id).filter(
            Conversation.constitution_type == ctype
        ).distinct().all()
        cust_id_list = [c[0] for c in cust_ids]
        order_count = db.query(Order).filter(
            Order.customer_id.in_(cust_id_list),
            Order.status == "paid"
        ).count() if cust_id_list else 0

        result.append({
            "constitution_type": ctype,
            "conversation_count": count,
            "order_count": order_count,
            "conversion_rate": round(order_count / count * 100, 1) if count else 0,
        })

    result.sort(key=lambda x: -x["conversation_count"])
    return {"data": result}


@router.patch("/conversations/{conv_id}/flag")
def flag_conversation(conv_id: int, db=Depends(get_db)):
    """切换对话的标记状态"""
    conv = db.query(Conversation).filter(Conversation.id == conv_id).first()
    if not conv:
        raise HTTPException(status_code=404, detail="对话不存在")
    conv.is_flagged = not conv.is_flagged
    db.commit()
    return {"id": conv_id, "is_flagged": conv.is_flagged}


@router.patch("/conversations/{conv_id}/notes")
def update_conversation_notes(conv_id: int, data: dict = Body(...), db=Depends(get_db)):
    """更新店员备注和标签"""
    conv = db.query(Conversation).filter(Conversation.id == conv_id).first()
    if not conv:
        raise HTTPException(status_code=404, detail="对话不存在")
    if "staff_notes" in data:
        conv.staff_notes = data["staff_notes"]
    if "staff_tags" in data:
        conv.staff_tags = data["staff_tags"]
    db.commit()
    return {"id": conv_id, "staff_notes": conv.staff_notes, "staff_tags": conv.staff_tags}


# ── LLM Review ──────────────────────────────────────────────────────────────

@router.post("/reviews/run")
def run_review(days: int = 1, db=Depends(get_db)):
    """触发LLM审核近N天对话"""
    from app.services.review_service import ReviewService
    svc = ReviewService()
    result = svc.review_conversations(db, days=days)
    return result


@router.get("/reviews")
def list_reviews(page: int = 1, page_size: int = 20, db=Depends(get_db)):
    """列出审核记录"""
    from app.models.review_result import ReviewResult
    total = db.query(ReviewResult).count()
    offset = (page - 1) * page_size
    reviews = db.query(ReviewResult).order_by(
        ReviewResult.reviewed_at.desc()
    ).offset(offset).limit(page_size).all()
    return {
        "total": total,
        "reviews": [
            {
                "id": r.id,
                "conversation_id": r.conversation_id,
                "problems_found": json.loads(r.problems_found) if r.problems_found else [],
                "suggestions": json.loads(r.suggestions) if r.suggestions else [],
                "quality_score": r.quality_score,
                "reviewed_at": str(r.reviewed_at),
            }
            for r in reviews
        ]
    }


@router.get("/reviews/runs")
def list_review_runs(db=Depends(get_db)):
    """列出审核批次（按时间分组，去重）"""
    from app.models.review_result import ReviewResult
    from sqlalchemy import func as sqlfunc, desc
    # Get unique review timestamps (group by minute)
    rows = db.query(
        sqlfunc.substr(ReviewResult.reviewed_at, 1, 16).label('batch'),
        sqlfunc.count(ReviewResult.id).label('count'),
        sqlfunc.max(ReviewResult.quality_score).label('score'),
    ).group_by('batch').order_by(desc('batch')).limit(20).all()

    runs = []
    for batch, count, score in rows:
        # Get one record from this batch to extract problems/suggestions
        sample = db.query(ReviewResult).filter(
            ReviewResult.reviewed_at.like(batch + '%')
        ).first()
        runs.append({
            "batch": batch,
            "conversation_count": count,
            "quality_score": score,
            "problems_found": json.loads(sample.problems_found) if sample and sample.problems_found else [],
            "suggestions": json.loads(sample.suggestions) if sample and sample.suggestions else [],
        })
    return {"runs": runs}