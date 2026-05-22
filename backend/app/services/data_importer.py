import openpyxl
import re
from pathlib import Path
from app.models.product import Product

PRODUCT_DIR = Path("E:/VIBE/beicaoji/beicaoji-产品目录")
FILE_MAP = {"biscuit": "biscuit.xlsx", "bread": "bread.xlsx", "tea": "tea.xlsx", "toy": "toy.xlsx"}
# Default prices by category (can be overridden when Excel includes price column)
CATEGORY_DEFAULT_PRICE = {"biscuit": 29, "bread": 25, "tea": 39, "toy": 19}

FIELD_ALIASES = {
    "sku_id": ["条码", "商品条码", "sku_id", "SKU_ID", "商品编码"],
    "name": ["商品名称", "名称", "name", "品名"],
    "price": ["零售价", "价格", "price", "售价"],
    "category": ["一级分类", "分类", "category", "商品分类"],
    "ingredients": ["成分", "ingredients", "配料", "主要成分"],
    "feature_tag": ["商品标签", "feature_tag", "标签"],
    "scene_tags": ["场景标签", "scene_tags", "适用场景"],
    "sales_script": ["销售话术", "sales_script", "话术"],
    "contraindication_tags": ["禁忌标签", "禁忌", "contraindication_tags"],
}


def parse_excel_columns(headers: list[str]) -> dict[str, int]:
    """
    Map header names to field names dynamically.
    Returns dict of {field_name: column_index}.
    Handles case-insensitive and fuzzy matching.
    """
    result = {}
    for idx, header in enumerate(headers):
        if not header:
            continue
        header_clean = str(header).strip()
        for field_name, aliases in FIELD_ALIASES.items():
            for alias in aliases:
                if alias.lower() == header_clean.lower() or alias in header_clean:
                    if field_name not in result:
                        result[field_name] = idx
                    break
    return result


def import_all(session):
    count = 0
    for category, filename in FILE_MAP.items():
        count += import_sheet(session, PRODUCT_DIR / filename, category)
    session.commit()
    return count


def import_sheet(session, filepath, category):
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]
    count = 0
    for row in list(ws.iter_rows(min_row=2, values_only=True)):
        if not row[0]:
            continue
        product = Product(
            sku_id=str(row[0]).strip(),
            category=category,
            feature_tag=str(row[1] or "").strip(),
            name=str(row[2] or "").strip(),
            ingredients=str(row[3] or "").strip(),
            scene_tags=str(row[4] or "").strip(),
            sales_script=str(row[5] or "").strip(),
            contraindication_tags=str(row[6] or "").strip(),
            price=CATEGORY_DEFAULT_PRICE.get(category, 25),
        )
        session.add(product)
        count += 1
    return count